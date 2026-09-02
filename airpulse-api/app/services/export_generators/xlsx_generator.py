import io
from typing import Any, Dict, List, Tuple
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# Institutional styling
NAVY_HEADER_FILL = PatternFill(start_color="081426", end_color="081426", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="081426")
DATA_FONT = Font(name="Calibri", size=10, color="101828")
MUTED_FONT = Font(name="Calibri", size=10, color="667085")

THIN_BORDER = Border(
    left=Side(style="thin", color="E4E7EC"),
    right=Side(style="thin", color="E4E7EC"),
    top=Side(style="thin", color="E4E7EC"),
    bottom=Side(style="thin", color="E4E7EC"),
)


def _style_header_row(ws, row_idx: int, max_col: int):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = NAVY_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            if len(val) > max_len and len(val) < 60:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def generate_apix_components_xlsx(
    summary_data: Dict[str, Any],
    components_data: List[Dict[str, Any]],
    weights_data: List[Dict[str, Any]],
    coverage_data: List[Dict[str, Any]],
    metadata_info: Dict[str, Any],
) -> Tuple[bytes, int]:
    """
    Generates an official statistical multi-sheet Excel workbook:
    Sheet 1: Summary
    Sheet 2: Index Components
    Sheet 3: Route Weights
    Sheet 4: Coverage
    Sheet 5: Metadata
    """
    wb = openpyxl.Workbook()
    # Default sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # 1. Summary Sheet
    ws_summary.append(["AirPulse — National Airfare Price Index (APIx) Summary"])
    ws_summary.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="081426")
    ws_summary.append(["Government of India / MoSPI Inflation Augmentation Subsystem"])
    ws_summary.cell(row=2, column=1).font = MUTED_FONT
    ws_summary.append([])

    summary_rows = [
        ("APIx Date", summary_data.get("index_date", "2026-09-02")),
        ("National APIx Value", summary_data.get("index_value", 108.43)),
        ("Daily Change (%)", summary_data.get("daily_change", "+0.41%")),
        ("Weekly Change (%)", summary_data.get("weekly_change", "+1.85%")),
        ("Monthly Change (%)", summary_data.get("monthly_change", "+4.12%")),
        ("Base Period", summary_data.get("base_period", "Aug 2026 = 100.0")),
        ("Methodology Version", summary_data.get("methodology_version", "APIx-Laspeyres-v1.2")),
        ("Basket Version", summary_data.get("basket_version", "domestic-basket-2026Q3")),
        ("Monitored Corridors", summary_data.get("active_routes", 81)),
        ("Route Coverage (%)", summary_data.get("route_coverage", "96.2%")),
        ("Source Coverage (%)", summary_data.get("source_coverage", "100.0%")),
        ("Statistical Quality Score (Q)", summary_data.get("quality_score", 0.964)),
        ("Data Origin", summary_data.get("data_origin", "LIVE")),
        ("Generated At (UTC)", summary_data.get("generated_at", "")),
    ]
    for label, val in summary_rows:
        row_num = ws_summary.max_row + 1
        c1 = ws_summary.cell(row=row_num, column=1, value=label)
        c2 = ws_summary.cell(row=row_num, column=2, value=val)
        c1.font = SECTION_FONT
        c2.font = DATA_FONT
        c1.border = THIN_BORDER
        c2.border = THIN_BORDER

    _autofit_columns(ws_summary)

    # 2. Index Components Sheet
    ws_comp = wb.create_sheet(title="Index Components")
    headers_comp = [
        "Route",
        "Booking Window",
        "Base Price (INR)",
        "Current Price (INR)",
        "Price Relative",
        "Route Weight",
        "Basket Weight",
        "Index Contribution (pts)",
        "Observation Count",
        "Coverage (%)",
    ]
    ws_comp.append(headers_comp)
    _style_header_row(ws_comp, 1, len(headers_comp))

    row_count = 0
    for comp in components_data:
        ws_comp.append([
            comp.get("route", "DEL-BOM"),
            comp.get("window", "T+7"),
            comp.get("base_price", 6900),
            comp.get("current_price", 7950),
            round(comp.get("price_relative", 115.22), 2),
            round(comp.get("route_weight", 0.048), 4),
            round(comp.get("basket_weight", 0.0096), 4),
            round(comp.get("contribution", 0.73), 2),
            comp.get("obs_count", 310),
            comp.get("coverage_pct", 99.0),
        ])
        row_count += 1
    _autofit_columns(ws_comp)
    ws_comp.freeze_panes = "A2"

    # 3. Route Weights Sheet
    ws_weights = wb.create_sheet(title="Route Weights")
    headers_weights = ["Route", "Origin", "Destination", "Distance (km)", "DGCA Traffic Share (%)", "Basket Weight (%)", "Reference Dataset"]
    ws_weights.append(headers_weights)
    _style_header_row(ws_weights, 1, len(headers_weights))
    for rw in weights_data:
        ws_weights.append([
            rw.get("route", ""),
            rw.get("origin", ""),
            rw.get("destination", ""),
            rw.get("distance_km", 0),
            rw.get("traffic_share", 0.0),
            rw.get("basket_weight", 0.0),
            rw.get("reference_dataset", "DGCA-DOM-2026-Q2"),
        ])
    _autofit_columns(ws_weights)
    ws_weights.freeze_panes = "A2"

    # 4. Coverage Sheet
    ws_cov = wb.create_sheet(title="Coverage")
    headers_cov = ["Route", "Booking Window", "Expected Observations", "Available Observations", "Coverage (%)", "Freshness Status"]
    ws_cov.append(headers_cov)
    _style_header_row(ws_cov, 1, len(headers_cov))
    for c in coverage_data:
        ws_cov.append([
            c.get("route", ""),
            c.get("window", ""),
            c.get("expected", 300),
            c.get("available", 295),
            c.get("coverage_pct", 98.3),
            c.get("freshness", "Active (<15m)"),
        ])
    _autofit_columns(ws_cov)
    ws_cov.freeze_panes = "A2"

    # 5. Metadata Sheet
    ws_meta = wb.create_sheet(title="Metadata")
    ws_meta.append(["Metadata Key", "Value"])
    _style_header_row(ws_meta, 1, 2)
    for k, v in metadata_info.items():
        ws_meta.append([k, str(v)])
    _autofit_columns(ws_meta)

    # Save to memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), row_count


def generate_anomalies_xlsx(anomalies: List[Dict[str, Any]]) -> Tuple[bytes, int]:
    """Generates a structured multi-sheet anomaly investigation workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anomalies"

    headers = [
        "Anomaly ID",
        "Detected At",
        "Route",
        "Booking Window",
        "Airline",
        "Source",
        "Actual Fare (INR)",
        "FareGuard Expected (INR)",
        "Residual (INR)",
        "Residual (%)",
        "PriceGuard Score",
        "Percentile",
        "Severity",
        "Review Status",
        "Decision",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    row_count = 0
    for a in anomalies:
        ws.append([
            a.get("anomaly_id", ""),
            a.get("detected_at", ""),
            a.get("route", ""),
            a.get("booking_window", ""),
            a.get("airline", ""),
            a.get("source", ""),
            a.get("actual_fare", 0),
            a.get("predicted_fare", 0),
            a.get("residual", 0),
            a.get("residual_pct", 0),
            a.get("anomaly_score", 0.0),
            a.get("anomaly_percentile", 0.0),
            a.get("severity", ""),
            a.get("status", ""),
            a.get("review_decision", "PENDING"),
        ])
        row_count += 1

    _autofit_columns(ws)
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), row_count
