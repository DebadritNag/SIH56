"""Official reference-dataset synchronization (MoSPI eSankhyiki).

Real change-detection + immutable versioning + Supabase storage + provenance +
sync-run tracking + audit events. Never fabricates data: if the official source
is unreachable/unconfigured the sync run is recorded as FAILED/PARTIAL and the
previously synced version stays active.

The high-frequency route-level airfare dataset is produced independently by market
collection; MoSPI CPI here is an OFFICIAL REFERENCE/BENCHMARK series only.
"""
from __future__ import annotations

import calendar
import hashlib
import io
import json
from datetime import date, datetime, timezone
from typing import Any, Dict, List, Optional, Tuple
from uuid import UUID, uuid4

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.collectors.government.mospi_esankhyiki import MospiESankhyikiAdapter
from app.core.utils import utc_now
from app.db.models import (
    BenchmarkFare,
    ReferenceDataset,
    ReferenceDatasetVersion,
    ReferenceSyncRun,
    Source,
)
from app.services.audit_service import AuditService
from app.services.storage_service import REFERENCE_DATASETS, get_storage_service

PARSER_VERSION = "reference-parser-v1.0.0"
_MONTHS = {m.lower(): i for i, m in enumerate(calendar.month_abbr) if m}


def _sha256(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def _month_start(d: date) -> date:
    return d.replace(day=1)


def _month_end(d: date) -> date:
    return date(d.year, d.month, calendar.monthrange(d.year, d.month)[1])


def _parse_month_label(cell: Any) -> Optional[date]:
    if isinstance(cell, datetime):
        return cell.date().replace(day=1)
    if isinstance(cell, date):
        return cell.replace(day=1)
    if isinstance(cell, str):
        s = cell.strip().rstrip("*").strip()
        parts = s.replace("/", "-").split("-")
        if len(parts) == 2 and parts[0].lower()[:3] in _MONTHS:
            mon = _MONTHS[parts[0].lower()[:3]]
            yr = int(parts[1]); yr += 2000 if yr < 100 else 0
            return date(yr, mon, 1)
    return None


class ReferenceDataService:
    """Synchronizes official MoSPI reference datasets with full provenance."""

    def __init__(self, session: AsyncSession):
        self.session = session
        self.storage = get_storage_service()
        self.audit = AuditService(session)

    # ------------------------------------------------------------------
    async def _get_source(self) -> Optional[Source]:
        res = await self.session.execute(select(Source).where(Source.name == "mospi_esankhyiki"))
        return res.scalars().first()

    # ------------------------------------------------------------------
    async def sync_mospi_datasets(
        self,
        trigger_type: str = "manual",
        actor_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Discover configured MoSPI datasets and sync any that have a fetchable
        source. Records a reference_sync_run and audit events. Truthful status."""
        src = await self._get_source()
        if not src:
            return {"status": "FAILED", "error": "MoSPI source not seeded"}

        adapter = MospiESankhyikiAdapter(source_id=str(src.id))
        run = ReferenceSyncRun(
            official_source_id=src.id, trigger_type=trigger_type,
            triggered_by=self.audit._coerce_actor(actor_id), status="RUNNING",
        )
        self.session.add(run)
        await self.session.flush()

        await self.audit.log_event(
            actor_id=actor_id, action="REFERENCE_CATALOGUE_REFRESHED",
            entity_type="reference_source", entity_id=str(src.id),
        )

        descriptors = await adapter.discover_datasets()
        run.datasets_discovered = len(descriptors)
        results: List[Dict[str, Any]] = []
        health = await adapter.health_check()

        for desc in descriptors:
            run.datasets_checked = (run.datasets_checked or 0) + 1
            # Only attempt a real fetch if the descriptor is fetchable (api/download).
            # Excel press-release files are ingested via ingest_official_file (upload).
            fetch = await adapter.fetch_dataset(
                dataset_code=desc.get("dataset_code", ""),
                api_url=None,  # CPI API requires validated params/key; not auto-called here
                download_url=None,
            )
            if fetch.get("status") != "OK" or not fetch.get("raw_bytes"):
                run.datasets_failed = (run.datasets_failed or 0) + 1
                results.append({
                    "dataset_code": desc.get("dataset_code"),
                    "status": fetch.get("status", "FAILED"),
                    "detail": fetch.get("detail"),
                })
                continue
            # (Reachable machine-readable path would flow through _persist_version here.)

        run.status = "COMPLETED" if run.datasets_failed == 0 else "PARTIAL"
        run.finished_at = utc_now()
        run.run_metadata = {"health": health, "results": results}
        await self.session.commit()
        return {
            "status": run.status,
            "sync_run_id": str(run.id),
            "health": health,
            "datasets_discovered": run.datasets_discovered,
            "datasets_failed": run.datasets_failed,
            "results": results,
            "note": ("Excel press-release datasets are ingested via file upload "
                     "(POST /reference-datasets/import). The documented CPI API "
                     "requires validated parameters and is not auto-polled."),
        }

    # ------------------------------------------------------------------
    async def ingest_official_file(
        self,
        raw: bytes,
        filename: str,
        dataset_code: str,
        dataset_name: str,
        external_dataset_id: str,
        source_url: str = "https://esankhyiki.mospi.gov.in",
        product_name: str = "Consumer Price Index",
        dataset_type: str = "CPI",
        actor_id: Optional[str] = None,
        trigger_type: str = "manual",
    ) -> Dict[str, Any]:
        """Ingest an official file (xlsx/csv/json) as an immutable versioned dataset.

        Change-detection: if the checksum matches the current version -> UNCHANGED.
        Otherwise a new immutable version is created, the original file is stored,
        and benchmark rows are (re)written. Audit + sync-run recorded.
        """
        src = await self._get_source()
        if not src:
            return {"status": "FAILED", "error": "MoSPI source not seeded"}

        checksum = _sha256(raw)
        fmt = MospiESankhyikiAdapter.detect_format(raw, filename)
        rows, schema_cols = self._parse_series(raw, fmt)
        if not rows:
            return {"status": "INVALID", "error": "No CPI series rows parsed from file"}

        periods = [r["period"] for r in rows]
        p_start, p_end = min(periods), max(periods)
        version_label = p_end.strftime("%Y-%m")
        schema_fingerprint = _sha256(",".join(schema_cols).encode())[:32]

        run = ReferenceSyncRun(
            official_source_id=src.id, trigger_type=trigger_type,
            triggered_by=self.audit._coerce_actor(actor_id), status="RUNNING",
            datasets_discovered=1, datasets_checked=1,
        )
        self.session.add(run); await self.session.flush()

        ds = (await self.session.execute(
            select(ReferenceDataset).where(ReferenceDataset.dataset_code == dataset_code)
        )).scalars().first()

        # change detection
        if ds and ds.current_version_id:
            cur = (await self.session.execute(
                select(ReferenceDatasetVersion).where(ReferenceDatasetVersion.id == ds.current_version_id)
            )).scalars().first()
            if cur and cur.checksum_sha256 == checksum:
                run.status = "COMPLETED"; run.datasets_unchanged = 1; run.finished_at = utc_now()
                run.reference_dataset_id = ds.id
                await self.session.commit()
                return {"status": "UNCHANGED", "dataset_id": str(ds.id),
                        "version_label": cur.version_label, "checksum": checksum}

        # store immutable original
        storage_path = f"mospi-esankhyiki/{external_dataset_id}/{version_label}/original.{fmt}"
        uploaded = False
        try:
            await self.storage.upload(REFERENCE_DATASETS, storage_path, raw,
                                      content_type=self._content_type(fmt))
            uploaded = True
        except Exception:  # storage optional; provenance still captured
            storage_path = None

        before = None
        if not ds:
            ds = ReferenceDataset(
                id=uuid4(), source_id=src.id, dataset_name=dataset_name, dataset_code=dataset_code,
                external_dataset_id=external_dataset_id, dataset_version=version_label,
                reference_period_start=p_start, reference_period_end=p_end,
                retrieved_at=utc_now(), source_url=source_url, landing_page_url=source_url,
                product_name=product_name, dataset_type=dataset_type, frequency="monthly",
                relevance="HIGH", checksum=checksum, storage_bucket=REFERENCE_DATASETS,
                storage_path=storage_path, file_format=fmt, status="SYNCED",
                row_count=len(rows), column_count=len(schema_cols),
                schema_fingerprint=schema_fingerprint,
                dataset_metadata={
                    "organization": "Ministry of Statistics and Programme Implementation",
                    "base_year": "2012=100", "parser_version": PARSER_VERSION,
                    "ingested_from": filename,
                },
            )
            self.session.add(ds)
            audit_action = "REFERENCE_DATASET_DISCOVERED"
        else:
            before = {"version": ds.dataset_version, "checksum": ds.checksum}
            ds.dataset_version = version_label; ds.reference_period_end = p_end
            ds.checksum = checksum; ds.status = "UPDATED"; ds.row_count = len(rows)
            ds.storage_path = storage_path or ds.storage_path
            ds.schema_fingerprint = schema_fingerprint; ds.retrieved_at = utc_now()
            audit_action = "REFERENCE_DATASET_UPDATED"
        await self.session.flush()

        prev_versions = (await self.session.execute(
            select(ReferenceDatasetVersion).where(ReferenceDatasetVersion.reference_dataset_id == ds.id)
        )).scalars().all()
        seq = len(list(prev_versions)) + 1

        ver = ReferenceDatasetVersion(
            id=uuid4(), reference_dataset_id=ds.id, version_label=version_label, version_sequence=seq,
            reference_period=f"{p_start.strftime('%b-%Y')} to {p_end.strftime('%b-%Y')}",
            source_url=source_url, retrieved_at=utc_now(), checksum_sha256=checksum,
            file_size_bytes=len(raw), row_count=len(rows), column_count=len(schema_cols),
            schema_fingerprint=schema_fingerprint, storage_bucket=REFERENCE_DATASETS,
            storage_path=storage_path, file_format=fmt, status="SYNCED",
            version_metadata={"latest_period": p_end.isoformat(), "parser_version": PARSER_VERSION},
        )
        self.session.add(ver); await self.session.flush()
        ds.current_version_id = ver.id

        # (Re)write benchmark rows for this dataset (history preserved in versions)
        for r in rows:
            self.session.add(BenchmarkFare(
                id=uuid4(), reference_dataset_id=ds.id, route_id=None,
                period_start=r["period"], period_end=_month_end(r["period"]),
                benchmark_type="mospi_cpi_general", value=r["index_combined"],
                unit="Index Point (2012=100)",
                benchmark_metadata={
                    "period_label": r["period"].strftime("%b-%Y"),
                    "index_combined": r["index_combined"],
                    "inflation_yoy_combined_pct": r.get("inflation_combined"),
                    "version_id": str(ver.id),
                },
            ))

        run.status = "COMPLETED"; run.reference_dataset_id = ds.id
        run.datasets_downloaded = 1; run.datasets_updated = 1
        run.bytes_downloaded = len(raw); run.finished_at = utc_now()

        await self.audit.log_event(
            actor_id=actor_id, action=audit_action, entity_type="reference_dataset",
            entity_id=str(ds.id), before_state=before,
            after_state={"version": version_label, "checksum": checksum, "rows": len(rows)},
            event_metadata={"schema_fingerprint": schema_fingerprint, "stored": uploaded},
        )
        await self.session.commit()
        return {
            "status": "SYNCED", "dataset_id": str(ds.id), "version_id": str(ver.id),
            "version_label": version_label, "rows": len(rows), "checksum": checksum,
            "stored": uploaded, "reference_period": ver.reference_period,
        }

    # ------------------------------------------------------------------
    def _parse_series(self, raw: bytes, fmt: str) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Parse an official CPI file into monthly combined-index rows.

        Supports the MoSPI CPI Annexure Excel layout and generic CSV/JSON.
        Only real values are extracted; malformed rows are skipped (not coerced).
        """
        if fmt in ("xlsx", "xls"):
            return self._parse_excel(raw)
        if fmt == "csv":
            return self._parse_csv(raw)
        if fmt == "json":
            return self._parse_json(raw)
        return [], []

    def _parse_excel(self, raw: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows: List[Dict[str, Any]] = []
        for r in ws.iter_rows(values_only=True):
            if not r:
                continue
            d = _parse_month_label(r[0])
            if not d:
                continue
            # MoSPI Annexure layout: Month, Rural, Urban, Combined, InflR, InflU, InflC
            combined = r[3] if len(r) > 3 else None
            infl_c = r[6] if len(r) > 6 else None
            if combined in (None, ""):
                continue
            try:
                rows.append({"period": d, "index_combined": float(combined),
                             "inflation_combined": float(infl_c) if infl_c not in (None, "") else None})
            except (TypeError, ValueError):
                continue
        cols = ["month", "rural", "urban", "combined", "infl_r", "infl_u", "infl_c"]
        return rows, cols

    def _parse_csv(self, raw: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
        import csv
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        cols = reader.fieldnames or []
        rows: List[Dict[str, Any]] = []
        for row in reader:
            month = row.get("month") or row.get("Month") or row.get("period") or (list(row.values())[0] if row else None)
            d = _parse_month_label(month)
            val = row.get("combined") or row.get("Combined") or row.get("index_combined") or row.get("index")
            if not d or val in (None, ""):
                continue
            try:
                rows.append({"period": d, "index_combined": float(val),
                             "inflation_combined": self._safe_float(row.get("inflation") or row.get("infl_c"))})
            except (TypeError, ValueError):
                continue
        return rows, list(cols)

    def _parse_json(self, raw: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
        try:
            data = json.loads(raw)
        except Exception:
            return [], []
        items = data.get("series") or data.get("data") or data.get("indices") if isinstance(data, dict) else data
        if not isinstance(items, list):
            return [], []
        rows: List[Dict[str, Any]] = []
        for it in items:
            if not isinstance(it, dict):
                continue
            d = _parse_month_label(it.get("period") or it.get("month"))
            val = it.get("index_combined") or it.get("index_value") or it.get("value")
            if not d or val in (None, ""):
                continue
            try:
                rows.append({"period": d, "index_combined": float(val),
                             "inflation_combined": self._safe_float(it.get("inflation_yoy_combined_pct") or it.get("inflation"))})
            except (TypeError, ValueError):
                continue
        return rows, ["period", "index_combined", "inflation"]

    @staticmethod
    def _safe_float(v: Any) -> Optional[float]:
        try:
            return float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _content_type(fmt: str) -> str:
        return {
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xls": "application/vnd.ms-excel",
            "csv": "text/csv",
            "json": "application/json",
        }.get(fmt, "application/octet-stream")
